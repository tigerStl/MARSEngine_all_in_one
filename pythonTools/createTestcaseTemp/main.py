import json
import os
from datetime import datetime

from openpyxl import load_workbook


SOURCE_DIR = os.path.join(os.path.dirname(__file__), "sourceExportedFile")


def _safe_cell_value(value):
    if value is None:
        return ""
    return value


def _build_steps(sheet):
    keys = [sheet.cell(row=6, column=col).value for col in range(1, 7)]
    keys = [_safe_cell_value(v) for v in keys]
    if len(keys) >= 5:
        keys[4] = "IsSkip"
    if len(keys) >= 6:
        keys[5] = "Data"

    steps = []
    run_order = 1
    row = 9
    while True:
        row_values = [_safe_cell_value(sheet.cell(row=row, column=col).value) for col in range(1, 7)]
        if all(v == "" for v in row_values):
            break

        step = {keys[i]: row_values[i] for i in range(6)}
        step["RunOrder"] = run_order
        step["MessageWhenGenByAI"] = ""
        steps.append(step)

        run_order += 1
        row += 1

    return steps


def _convert_workbook(xlsx_path):
    workbook = load_workbook(xlsx_path, data_only=True)
    sheet = workbook.worksheets[0]

    steps = _build_steps(sheet)
    base_name = os.path.splitext(os.path.basename(xlsx_path))[0]
    payload = {
        "name": base_name,
        "description": "MARS AI auto generate test case temp",
        "version": "1",
        "TestSteps": steps,
        "CreateDate": datetime.now().isoformat(sep=" ", timespec="seconds"),
        "Creator": "pythonUtility",
    }

    json_path = os.path.splitext(xlsx_path)[0] + ".json"
    with open(json_path, "w", encoding="utf-8") as handle:
        json.dump(payload, handle, ensure_ascii=False, indent=2)


def main():
    if not os.path.isdir(SOURCE_DIR):
        raise FileNotFoundError(f"sourceExportedFile not found: {SOURCE_DIR}")

    for filename in os.listdir(SOURCE_DIR):
        if not filename.lower().endswith(".xlsx"):
            continue
        if filename.startswith("~$"):
            continue
        _convert_workbook(os.path.join(SOURCE_DIR, filename))


if __name__ == "__main__":
    main()
